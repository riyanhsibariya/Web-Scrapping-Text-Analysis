import os
import re
import nltk
import openpyxl
from textstat import syllable_count


# Function to load words from a text file
def load_words_from_file(filename):
    """Loads words from a text file, handling potential file I/O errors."""
    try:
        with open(filename, 'r') as file:
            words = set(file.read().splitlines())
    except FileNotFoundError:
        print(f"Error: File not found - {filename}")
        words = set()  # Create an empty set in case of error
    return words
def load_dictionaries():
    #Loading all stop words
    stop_words = load_words_from_file('StopWords_Auditor.txt')
    stop_words.update(load_words_from_file('StopWords_Currencies.txt'))
    stop_words.update(load_words_from_file('StopWords_DatesandNumbers.txt'))
    stop_words.update(load_words_from_file('StopWords_Generic.txt'))
    stop_words.update(load_words_from_file('StopWords_GenericLong.txt'))
    stop_words.update(load_words_from_file('StopWords_Geographic.txt'))
    stop_words.update(load_words_from_file('StopWords_Names.txt'))

    # Loading all positive and negative words
    positive_words = load_words_from_file('positive-words.txt') - stop_words  # Exclude stop words from positive
    negative_words = load_words_from_file('negative-words.txt') - stop_words  # Exclude stop words from negative
    return stop_words, positive_words, negative_words


# Function for sentiment analysis
def analyze_sentiment(text, stop_words, positive_words, negative_words):

    tokens = text.split()
    cleaned_tokens = [word.lower() for word in tokens if word not in stop_words]

    positive_score = sum(1 for word in cleaned_tokens if word in positive_words)
    negative_score = -1 * sum(-1 for word in cleaned_tokens if word in negative_words)  # Ensures positive value

    if positive_score + negative_score == 0:
        polarity_score = 0
        subjectivity_score = 0
    else:
        polarity_score = (positive_score - negative_score) / (positive_score + negative_score + 0.000001)
        subjectivity_score = (positive_score + negative_score) / (len(cleaned_tokens) + 0.000001)

    return positive_score, negative_score, polarity_score, subjectivity_score

# Function to calculate readability
def analyze_readability(text, positive_words):
    sentences = nltk.sent_tokenize(text)
    word_count = len(text.split())

    if not sentences:
        return 0, 0, 0  # Handle cases with no sentences

    avg_sentence_length = word_count / len(sentences)

    complex_words = sum(1 for word in text.split() if count_syllables(word) > 2)
    percentage_complex_words = (complex_words / word_count) * 100

    fog_index = 0.4 * (avg_sentence_length + percentage_complex_words)
    return avg_sentence_length, percentage_complex_words, fog_index


def count_syllables(word, positive_words=set()):
    vowels = 'aeiouAEIOU'
    syllables = 0
    if word.endswith('es') or word.endswith('ed') and word not in positive_words:
        return 1  # Handle special cases like "love" or "looked" (not positive/negative words)
    for idx, char in enumerate(word):
        if char in vowels and (idx == 0 or word[idx - 1] not in vowels):
            syllables += 1
    return syllables


# Function to count words
def count_words(text):
    """Counts the total number of words in the text."""
    return len(text.split())


# Function to find personal pronouns
def find_personal_pronouns(text):
    pronouns = r"I|we|my|ours|us"
    return len(re.findall(pronouns, text, flags=re.IGNORECASE))


# Function to calculate average word length
def calculate_avg_word_length(text):
    words = text.split()
    if not words:
        return 0
    return sum(len(word) for word in words) / len(words)

def analyze_text_file(filename, existing_excel_file, row):
    """Analyzes a text file and writes results to an existing Excel sheet"""
    wb = openpyxl.load_workbook(existing_excel_file)
    sheet = wb.active  # Use the active sheet (assuming you want to write to the current sheet)

    with open(f"extracted_texts/{filename}", "r", encoding="utf-8") as file:
        text = file.read()

    # Load dictionaries
    stop_words, positive_words, negative_words = load_dictionaries()

    # Perform analysis
    positive_score, negative_score, polarity_score, subjectivity_score = analyze_sentiment(text, stop_words, positive_words, negative_words)
    avg_sentence_length, percentage_complex_words, fog_index = analyze_readability(text, positive_words)
    word_count = count_words(text)
    syllables_per_word = sum(
        count_syllables(word, positive_words) for word in text.split()) / word_count if word_count else 0
    personal_pronouns = find_personal_pronouns(text)
    avg_word_length = calculate_avg_word_length(text)

    # Write data to specific columns starting from column 3 (as URL_ID and URL are already filled in columns 1 and 2)
    sheet.cell(row=row, column=3).value = positive_score
    sheet.cell(row=row, column=4).value = negative_score
    sheet.cell(row=row, column=5).value = polarity_score
    sheet.cell(row=row, column=6).value = subjectivity_score
    sheet.cell(row=row, column=7).value = avg_sentence_length
    sheet.cell(row=row, column=8).value = percentage_complex_words
    sheet.cell(row=row, column=9).value = fog_index
    sheet.cell(row=row, column=10).value = word_count / (len(nltk.sent_tokenize(text)) if text else 1)
    sheet.cell(row=row, column=11).value = sum(1 for word in text.split() if syllable_count(word) > 2)
    sheet.cell(row=row, column=12).value = word_count
    sheet.cell(row=row, column=13).value = syllables_per_word
    sheet.cell(row=row, column=14).value = personal_pronouns
    sheet.cell(row=row, column=15).value = avg_word_length

    wb.save(existing_excel_file)


row = 2  # Starts from row 2, as headers are in row 1
for filename in os.listdir("extracted_texts"):
    analyze_text_file(filename, "analysis_results.xlsx", row)
    row += 1  # Increments row number for the next analysis
